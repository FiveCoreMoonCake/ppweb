import openpyxl
import json

wb = openpyxl.load_workbook(r'D:\self_work\workspace\ppweb\testdata\地产权与未来利益全配对通关表.xlsx')
ws = wb.active

result = []
result.append(f"Sheet: {ws.title}")
result.append(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
result.append(f"Merged cells: {[str(m) for m in ws.merged_cells.ranges]}")
result.append("=" * 80)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    row_data = {}
    for cell in row:
        val = cell.value if cell.value is not None else ""
        row_data[cell.coordinate] = str(val)
    result.append(json.dumps(row_data, ensure_ascii=False))

with open(r'D:\self_work\workspace\ppweb\testdata\output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(result))

print("Done writing output.txt")
